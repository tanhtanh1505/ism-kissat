from openpyxl import Workbook
import random
import math
import pandas as pd
import matplotlib.pyplot as plt
import os

def extract_numbers(line):
    numbers = []
    is_end = False
    for word in line.split():
        if word.isdigit() or (word.startswith('-') and word[1:].isdigit()):
            numbers.append(int(word))
            if int(word) == 0:
                is_end = True
                break
    return numbers, is_end

def extract_solutions_from_result(filename):
    equation = []
    try:
        with open(filename, 'r') as file:
            found_result = False
            for line in file:
                if "c ---- [ result ]" in line:
                    found_result = True
                if found_result and line.startswith("v"):
                        e, is_end = extract_numbers(line.strip())
                        equation.extend(e)
                if "c ---- [ profiling ]" in line:
                    break
    except FileNotFoundError:
        print("File not found.")
    except Exception as e:
        print("An error occurred:", str(e))
    return equation

def save_equation_to_file(filename, equation, num_items, num_transactions, min_support):
    with open(filename, 'a') as f:
        item_set = set()
        valid_transactions = set()
        for item in equation:
            if item > 0:
                if item <= num_items:
                    item_set.add(item)
                elif item <= num_items + num_transactions:
                    valid_transactions.add(item - num_items)
        f.write("=================================================================\n")
        f.write(' '.join([str(x) for x in equation]) + '\n')
        f.write("Number of items: " + str(num_items) + '\n')
        f.write("Number of transactions: " + str(num_transactions) + '\n')
        f.write("Minimum support: " + str(min_support) + '\n')
        f.write("Item set found: " + str(item_set) + '\n')
        f.write("Valid transactions: " + str(valid_transactions) + '\n')
        f.write("=================================================================\n")

def ignore_solved_solutions(file_input, equation, n_items):
    # convert all to negative
    equation_for_items = []
    for x in range(n_items):
        equation_for_items.append(equation[x] * -1)
    equation_for_items.append(0)

    num_clauses = 0
    num_vals = 0
    # append all to file_input
    with open(file_input, 'a') as f:
        f.write(' '.join([str(x) for x in equation_for_items]) + '\n')
    # update the number of clauses
    with open(file_input, 'r') as f:
        lines = f.readlines()
        _,_, num_vals,num_clauses = lines[0].split()
        new_num_clauses = int(num_clauses) + 1
        lines[0] = "p cnf " + num_vals + " " + str(new_num_clauses) + "\n"
    # write back to file_input
    with open(file_input, 'w') as f:
        f.writelines(lines)

def get_max_item(clauses):
    max_item = 0
    for clause in clauses:
        for item in clause:
            max_item = max(max_item, abs(item))
    return max_item

def write_cnf_to_file(vars, clauses, output_file):
    with open(output_file, 'w') as writer:
        # Write a line of information about the number of variables and constraints
        writer.write("p cnf " + str(vars) + " " + str(len(clauses)) + "\n")
        # Write each clause to the file
        for clause in clauses:
            for literal in clause:
                writer.write(str(literal) + " ")
            writer.write("0\n")

def write_data_to_excel(data, path, is_raw = True):
    if is_raw:
        # Generate header
        raw_header = data[0].keys()
        header = []
        for key in raw_header:
            header.append(key)
        # write to excel
        book = Workbook()
        sheet = book.active
        sheet.append(header)
        for d in data:
            sheet.append([d[key] for key in raw_header])
        book.save(path)
        return
    
    # Generate header
    raw_header = data[0].keys()
    header = []
    sub_header = []
    list_need_merge = []
    cur_col = 'A'
    header_merge = ""
    start_merge_cel = ""
    end_merge_cel = ""
    for idx, key in enumerate(raw_header):
        num_sub_header = len(key.split("/"))
        if num_sub_header > 1:
            cur_header = key.split("/")[0]
            header.append(cur_header)
            sub_header.append(key.split("/")[1])
            
            if header_merge != cur_header:
                if start_merge_cel != "" and end_merge_cel != "" and start_merge_cel != end_merge_cel:
                    list_need_merge.append(f'{start_merge_cel}:{end_merge_cel}')
                start_merge_cel = cur_col + str(1)
                header_merge = cur_header
            
            end_merge_cel = cur_col + str(1)
            if idx == len(raw_header) - 1:
                list_need_merge.append(f'{start_merge_cel}:{end_merge_cel}')
        else:
            header.append(key)
            sub_header.append("")
            list_need_merge.append(f'{cur_col + str(1)}:{cur_col + str(2)}')
        cur_col = chr(ord(cur_col) + 1)
    # write to excel
    book = Workbook()
    sheet = book.active
    sheet.append(header)
    sheet.append(sub_header)
    for d in data:
        sheet.append([d[key] for key in raw_header])
    for merge in list_need_merge:
        sheet.merge_cells(merge)
    book.save(path)

def write_data_to_graph(data_from_path, output_path, modes = [1, 2, 3]):
    min_support_path = f'{output_path}/by_min_support'
    transactions_path = f'{output_path}/by_transactions'
    
    # create output path if not exist
    os.makedirs(min_support_path, exist_ok=True)
    os.makedirs(transactions_path, exist_ok=True)

    # read data from excel
    df = pd.read_excel(data_from_path)
    for n_items in df["num_items"].unique():
        for n_transactions in df["num_transactions"].unique():
            # create a single figure with subplots for each combination of n_items and n_transactions
            fig, axs = plt.subplots(1, 3, figsize=(18, 5))
            fig.suptitle(f"Number of items: {n_items}, Number of transactions: {n_transactions}")

            df_filtered = df[(df["num_items"] == n_items) & (df["num_transactions"] == n_transactions)]

            # subplot 1: number of clauses
            if 1 in modes:
                axs[0].plot(df_filtered["min_support"], df_filtered["standard/clauses"], label="Standard", color="red")
            if 3 in modes:
                axs[0].plot(df_filtered["min_support"], df_filtered["old_sequential_encoding/clauses"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                axs[0].plot(df_filtered["min_support"], df_filtered["sequential_encoding/clauses"], label="New Sequential Counter", color="green")
            axs[0].set_xlabel("Minimum support")
            axs[0].set_ylabel("Number of clauses")
            axs[0].legend()

            # subplot 2: time
            if 1 in modes:
                axs[1].plot(df_filtered["min_support"], df_filtered["standard/time"], label="Standard", color="red")
            if 3 in modes:
                axs[1].plot(df_filtered["min_support"], df_filtered["old_sequential_encoding/time"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                axs[1].plot(df_filtered["min_support"], df_filtered["sequential_encoding/time"], label="New Sequential Counter", color="green")
            axs[1].set_xlabel("Minimum support")
            axs[1].set_ylabel("Time")
            axs[1].legend()

            # subplot 3: number of variables
            if 1 in modes:
                axs[2].plot(df_filtered["min_support"], df_filtered["standard/vars"], label="Standard", color="red")
            if 3 in modes:
                axs[2].plot(df_filtered["min_support"], df_filtered["old_sequential_encoding/vars"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                axs[2].plot(df_filtered["min_support"], df_filtered["sequential_encoding/vars"], label="New Sequential Counter", color="green")
            axs[2].set_xlabel("Minimum support")
            axs[2].set_ylabel("Number of variables")
            axs[2].legend()

            plt.tight_layout(pad=3.0)
            plt.savefig(f'{min_support_path}/n_trans_{n_transactions}.png')
            plt.clf()

        for min_support in df["min_support"].unique():
            # create a single figure with subplots for each combination of n_items and min_support
            fig, axs = plt.subplots(1, 3, figsize=(18, 5))
            fig.suptitle(f"Number of items: {n_items}, Minimum support: {min_support}")

            df_filtered = df[(df["num_items"] == n_items) & (df["min_support"] == min_support)]

            # subplot 1: number of clauses
            if 1 in modes:
                axs[0].plot(df_filtered["num_transactions"], df_filtered["standard/clauses"], label="Standard", color="red")
            if 3 in modes:
                axs[0].plot(df_filtered["num_transactions"], df_filtered["old_sequential_encoding/clauses"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                axs[0].plot(df_filtered["num_transactions"], df_filtered["sequential_encoding/clauses"], label="New Sequential Counter", color="green")
            axs[0].set_xlabel("Number of transactions")
            axs[0].set_ylabel("Number of clauses")
            axs[0].legend()

            # subplot 2: time
            if 1 in modes:
                axs[1].plot(df_filtered["num_transactions"], df_filtered["standard/time"], label="Standard", color="red")
            if 3 in modes:    
                axs[1].plot(df_filtered["num_transactions"], df_filtered["old_sequential_encoding/time"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                axs[1].plot(df_filtered["num_transactions"], df_filtered["sequential_encoding/time"], label="New Sequential Counter", color="green")
            axs[1].set_xlabel("Number of transactions")
            axs[1].set_ylabel("Time")
            axs[1].legend()

            # subplot 3: number of variables
            if 1 in modes:
                axs[2].plot(df_filtered["num_transactions"], df_filtered["standard/vars"], label="Standard", color="red")
            if 3 in modes:
                axs[2].plot(df_filtered["num_transactions"], df_filtered["old_sequential_encoding/vars"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                axs[2].plot(df_filtered["num_transactions"], df_filtered["sequential_encoding/vars"], label="New Sequential Counter", color="green")
            axs[2].set_xlabel("Number of transactions")
            axs[2].set_ylabel("Number of variables")
            axs[2].legend()

            plt.tight_layout(pad=3.0)
            plt.savefig(f'{transactions_path}/min_supp_{min_support}.png')
            plt.clf()

def write_data_to_each_graph(data_from_path, output_path, modes = [1, 2, 3]):
    min_support_path = f'{output_path}/by_min_support'
    transactions_path = f'{output_path}/by_transactions'
    
    # create output path if not exist
    os.makedirs(min_support_path, exist_ok=True)
    os.makedirs(transactions_path, exist_ok=True)

    # read data from excel
    df = pd.read_excel(data_from_path)
    for n_items in df["num_items"].unique():
        for n_transactions in df["num_transactions"].unique():
            # create a single figure with subplots for each combination of n_items and n_transactions
            df_filtered = df[(df["num_items"] == n_items) & (df["num_transactions"] == n_transactions)]

            plt.figure(figsize=(6, 5))
            # subplot 1: number of clauses
            if 1 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["standard/clauses"], label="Standard", color="red")
            if 3 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["old_sequential_encoding/clauses"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["sequential_encoding/clauses"], label="New Sequential Counter", color="green")
            plt.xlabel("Minimum support")
            plt.ylabel("Number of clauses")
            plt.legend()

            plt.savefig(f'{min_support_path}/n_trans_{n_transactions}_clauses.png')
            plt.clf()
        
            # subplot 2: time
            if 1 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["standard/time"], label="Standard", color="red")
            if 3 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["old_sequential_encoding/time"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["sequential_encoding/time"], label="New Sequential Counter", color="green")
            plt.xlabel("Minimum support")
            plt.ylabel("Time")
            plt.legend()

            plt.savefig(f'{min_support_path}/n_trans_{n_transactions}_time.png')
            plt.clf()

            # subplot 3: number of variables
            if 1 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["standard/vars"], label="Standard", color="red")
            if 3 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["old_sequential_encoding/vars"], label="Old Sequential Counter", color="orange")
            if 2 in modes:
                plt.plot(df_filtered["min_support"], df_filtered["sequential_encoding/vars"], label="New Sequential Counter", color="green")
            plt.xlabel("Minimum support")
            plt.ylabel("Number of variables")
            plt.legend()

            plt.savefig(f'{min_support_path}/n_trans_{n_transactions}_vars.png')
            plt.clf()


def generate_input(n_items, n_transactions, output):
    with open(output, "w") as f:
        for i in range(n_transactions):
            for j in range(n_items):
                is_true = random.choice([0, 1])
                f.write(str(j*2 + is_true) + " ")
            f.write("\n")

def get_c_k_n(k, n):
    return int(math.factorial(n) / (math.factorial(k) * math.factorial(n - k)))

# write_data_to_graph("./output/benchmark.xlsx", "./output/20240505_22", [1, 2, 3])
write_data_to_each_graph("./output/benchmark.xlsx", "./output/20240505_22", [1, 2, 3])